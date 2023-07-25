# Import necessary modules
import json
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import numbers

# Define the JSON file names
files = ["1010.json", "1014.json", "2010.json", "2011.json", "3010.json", "3012.json"]

# Define a mapping from JSON file names to sheet names
sheet_names = {
    "1010.json": "برنامه‌سازی پیشرفته",
    "1014.json": "داده‌ساختارها و الگوریتم‌ها",
    "2010.json": "برنامه‌سازی پایتون",
    "2011.json": "ساختارهای گسسته",
    "3010.json": "برنامه‌سازی برای تحلیل داده",
    "3012.json": "ریاضیات هوش مصنوعی",
}

# Create a Pandas Excel writer using openpyxl as the engine
writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')

# Iterate over each JSON file
for file in files:
    # Open and load the JSON file
    with open(file) as f:
        json_data = json.load(f)

    # Extract data from JSON
    course_name = json_data["courses"][0]["name"]
    course_code = json_data["courses"][0]["code"]
    students = json_data["courses"][0]["current_group"][0]["students"]

    # Prepare the data for the DataFrame
    data = {
        "Course Name": [course_name] * len(students),
        "Course Code": [course_code] * len(students),
        "Student ID": [str(student["id"]) for student in students],
        "Name": [student["name"] for student in students],
        "Surname": [student["surname"] for student in students],
        "Gender": [student["gender"] for student in students],
        "Email": [student["email"] for student in students],
        "Mobile Number": [str(student["mobile_number"]) for student in students],
        "National Code": [str(student["national_code"]) for student in students],
        "Phone Number": [str(student["phone_number"]) for student in students],
        "Status": [student["pivot"]["status"] for student in students],
    }

    # Convert the data into a Pandas DataFrame
    df = pd.DataFrame(data)

    # Replace values in 'Status' column as per requirements
    df['Status'] = df['Status'].replace({'pending': 'ثبت‌نام', 'requesting': 'کردیت'})

    # Sort DataFrame by 'Status' column
    df = df.sort_values('Status')

    # Write the DataFrame to an Excel file
    df.to_excel(writer, sheet_name=sheet_names[file], index=False)  

# Save the changes in writer object
writer.save()

# Define Excel formatting variables
font = Font(name='Vazirmatn')
header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
other_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# Define column widths
column_widths = {"Email": 30, "Student ID": 10, "Course Code": 10, "Status": 10, "Gender": 10, "Name": 15, "Surname": 15}

# Load the workbook
book = openpyxl.load_workbook('output.xlsx')

# Iterate over each sheet in the workbook
for sheet in book.sheetnames:
    worksheet = book[sheet]
    # Iterate over each column in DataFrame and set column width and cell format
    for i, column in enumerate(df.columns, start=1):
        if column in column_widths:
            worksheet.column_dimensions[get_column_letter(i)].width = column_widths[column]
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = 20

        # if the column contains numbers, set the cell format to 'Text'
        if column in ["Student ID", "Mobile Number", "National Code", "Phone Number", "Status"]:
            for cell in worksheet[get_column_letter(i)]:
                cell.number_format = numbers.FORMAT_TEXT

    # Iterate over each cell in the worksheet and set alignment, font, and fill
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font
            if cell.row == 1:
                cell.fill = header_fill
            else:
                cell.fill = other_fill

# Save the changes in the workbook
book.save('output.xlsx')